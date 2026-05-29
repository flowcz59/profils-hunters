#!/usr/bin/env python3
"""
Recherche automatique de profils COBOL & Java - Région Nord (Hauts-de-France)
Auteur: Claude (Anthropic)
Usage: python recherche_profils_lille.py [--github-token TOKEN]

Note RGPD: Ce script collecte uniquement des données publiquement disponibles.
Les résultats doivent être traités conformément au RGPD (UE 2016/679).
"""

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import time
import sys
import os
import json
import argparse

try:
    try:
        from ddgs import DDGS
    except ImportError:
        from duckduckgo_search import DDGS
    HAS_DDG = True
except ImportError:
    HAS_DDG = False
    print("[AVERTISSEMENT] ddgs non disponible. Recherches web désactivées.")

# ─── Configuration ────────────────────────────────────────────────────────────

# ★ COLLE TON TOKEN GITHUB ICI (entre les guillemets) ★
GITHUB_TOKEN = "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"

TECHNOLOGIES = ["COBOL", "Java"]

# Région Nord étendue : Nord (59) + Pas-de-Calais (62) + termes génériques
LOCATIONS_NORD = [
    # Métropole Lilloise
    "Lille", "Roubaix", "Tourcoing", "Villeneuve-d'Ascq", "Armentières",
    "Lambersart", "Hellemmes", "Lomme", "Loos", "Seclin", "Wattignies",
    # Reste du Nord (59)
    "Valenciennes", "Douai", "Dunkerque", "Maubeuge", "Cambrai",
    "Denain", "Anzin", "Grande-Synthe", "Avesnes-sur-Helpe",
    # Pas-de-Calais (62)
    "Arras", "Lens", "Calais", "Boulogne-sur-Mer", "Béthune",
    "Liévin", "Hénin-Beaumont", "Saint-Omer", "Bruay-la-Buissière",
    # Termes régionaux (GitHub/SO les reconnaissent)
    "Hauts-de-France", "Nord-Pas-de-Calais", "Nord", "Pas-de-Calais",
    "Metropole Europeenne de Lille",
]

# Sous-ensemble utilisé pour les recherches GitHub (éviter rate-limit sans token)
LOCATIONS_GH_PRIORITY = [
    "Lille", "Roubaix", "Tourcoing", "Villeneuve-d'Ascq",
    "Valenciennes", "Douai", "Dunkerque", "Arras", "Lens",
    "Hauts-de-France", "Nord-Pas-de-Calais", "Nord",
]

# Villes pour filtrage des résultats SO
LOCATIONS_SO_FILTER = [l.lower() for l in LOCATIONS_NORD] + ["59", "62", "hauts de france", "nord pas de calais"]

# ─── Filtres qualité ──────────────────────────────────────────────────────────
# Entreprises à exclure (configurable)
EXCLUDED_COMPANIES = ["infotel"]

# LinkedIn subdomains = profils hors-France → exclus
NON_FR_LINKEDIN = [
    "in.linkedin.com", "ca.linkedin.com", "lt.linkedin.com", "au.linkedin.com",
    "sg.linkedin.com", "de.linkedin.com", "br.linkedin.com", "ng.linkedin.com",
    "pl.linkedin.com", "mc.linkedin.com", "ad.linkedin.com", "pt.linkedin.com",
    "mx.linkedin.com", "ar.linkedin.com", "pk.linkedin.com",
]

# Mots-clés bio indiquant clairement un profil hors-France
NON_FR_BIO_KEYWORDS = [
    "india", "bangalore", "hyderabad", "mumbai", "chennai", "delhi", "pune",
    "united states", "new york", "california", "texas", "chicago",
    "summit racing", "cumming, ga", "desjardins", "northeast regional",
    "singapore", "hong kong", "malaysia", "united kingdom", "london",
    "pakistan", "lahore", "karachi", "nigeria", "ghana", "south africa",
    "brazil", "argentina", "ukraine", "russia", "willemstad", "curaçao",
    "indira gandhi institute", "gokul institute", "vit vellore",
]

# Domaines = offres d'emploi ou contenus hors-sujet → exclus
BLOCKLIST_DOMAINS = [
    "hellowork.com", "indeed.", "glassdoor.", "francetravail.fr",
    "pole-emploi", "cadremploi.", "monster.", "welcometothejungle.",
    "jobteaser.", "meteojob.", "keljob.", "regionsjob.", "jobijoba.",
    "emplois-informatique.fr", "youtube.com", "google.com",
    "arxiv.org", "grouplens.org", "javalab.org", "ledgerlens.io",
    "docs.api.lens.org", "jesuisundev.com", "freelance-informatique.fr",
]
BLOCKLIST_URL_PATTERNS = [
    "/jobs/", "linkedin.com/jobs", "/emplois/", "/offre",
    "/recrutement", "offres-emploi", "q-emploi", "search?",
]
BLOCKLIST_NAME_PATTERNS = [
    "offres d'emploi", "recrutement :", "opportunités d'emploi",
]

def is_valid_profile(p: dict) -> bool:
    """Retourne False si le profil est une offre d'emploi, hors-sujet ou entreprise exclue."""
    url  = p.get("profil_url", "").lower()
    nom  = p.get("nom", "").lower()
    text = (nom + " " + p.get("entreprise", "") + " " + p.get("bio", "")).lower()

    if any(d in url for d in BLOCKLIST_DOMAINS):
        return False
    if any(pat in url for pat in BLOCKLIST_URL_PATTERNS):
        return False
    if any(pat in nom for pat in BLOCKLIST_NAME_PATTERNS):
        return False
    if any(company.lower() in text for company in EXCLUDED_COMPANIES):
        return False
    # Filtre géographique : LinkedIn sous-domaines étrangers
    if any(f"https://{nd}" in url or f"/{nd}/" in url for nd in NON_FR_LINKEDIN):
        return False
    # Filtre géographique : mots-clés bio indiquant hors-France
    bio_lower = (p.get("bio") or "").lower()
    if any(kw in bio_lower for kw in NON_FR_BIO_KEYWORDS):
        return False
    return True

OUTPUT_DIR     = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE    = os.path.join(OUTPUT_DIR, "profils_cobol_java_nord.xlsx")
SNAPSHOT_FILE  = os.path.join(OUTPUT_DIR, "snapshot_profils.json")   # URLs du run précédent
NOUVEAUX_FILE  = os.path.join(OUTPUT_DIR, "nouveaux_profils.xlsx")   # Rapport diff hebdo

# ★ CLÉ HUNTER.IO (optionnelle — plan gratuit : 25 recherches/mois) ★
# Créer un compte sur hunter.io pour obtenir une clé API gratuite
HUNTER_API_KEY = "1d57ab6e816524b0cb03dc43dfad07aaf0cf9bff"

HEADERS_GH = {"Accept": "application/vnd.github.v3+json"}

# ─── Couleurs charte ──────────────────────────────────────────────────────────
COLOR_HEADER     = "1F4E79"
COLOR_COBOL      = "D6E4F0"
COLOR_JAVA       = "D5F5E3"
COLOR_SOURCE_GH  = "EBF5FB"
COLOR_SOURCE_LI  = "EAF2FF"
COLOR_SOURCE_SO  = "FDF2E9"
COLOR_SOURCE_WEB = "F9F2FF"


# ══════════════════════════════════════════════════════════════════════════════
#  SOURCES
# ══════════════════════════════════════════════════════════════════════════════

def search_github(tech: str, github_token: str = "") -> list[dict]:
    """Recherche des utilisateurs GitHub avec COBOL ou Java dans la région lilloise."""
    profiles = []
    if github_token:
        HEADERS_GH["Authorization"] = f"token {github_token}"

    # Pour COBOL, les devs n'ont souvent pas de langage déclaré → chercher aussi par mot-clé bio
    if tech.upper() == "COBOL":
        queries_per_loc = [
            lambda loc: f"location:{loc} COBOL in:bio",
            lambda loc: f"location:{loc} language:COBOL",
        ]
    else:
        queries_per_loc = [
            lambda loc: f"location:{loc} language:{tech}",
        ]

    locs = LOCATIONS_GH_PRIORITY if not github_token else LOCATIONS_NORD
    for loc in locs:
        for q_fn in queries_per_loc:
            query = q_fn(loc)
            url = f"https://api.github.com/search/users?q={query}&per_page=30"
            try:
                r = requests.get(url, headers=HEADERS_GH, timeout=10)
                if r.status_code == 403:
                    print(f"  [GitHub] Rate limit atteint pour '{loc}'. Attendez 60s ou ajoutez un token.")
                    break
                if r.status_code != 200:
                    continue
                data = r.json()
                for user in data.get("items", []):
                    login = user.get("login", "")
                    detail_url = f"https://api.github.com/users/{login}"
                    detail_r = requests.get(detail_url, headers=HEADERS_GH, timeout=10)
                    if detail_r.status_code == 200:
                        d = detail_r.json()
                        profiles.append({
                            "nom":         d.get("name") or login,
                            "pseudo":      login,
                            "email":       d.get("email") or "Non public",
                            "entreprise":  d.get("company") or "",
                            "localisation": d.get("location") or loc,
                            "bio":         d.get("bio") or "",
                            "profil_url":  d.get("html_url", ""),
                            "repos":       d.get("public_repos", 0),
                            "followers":   d.get("followers", 0),
                            "disponibilite": "Indéterminée",
                            "technologie": tech,
                            "source":      "GitHub",
                        })
                    time.sleep(0.3)
            except Exception as e:
                print(f"  [GitHub] Erreur pour {loc}/{tech}: {e}")
            time.sleep(1)

    # Dédupliquer par pseudo
    seen = set()
    unique = []
    for p in profiles:
        key = p["pseudo"]
        if key not in seen:
            seen.add(key)
            unique.append(p)
    return unique


def search_stackoverflow(tech: str) -> list[dict]:
    """Recherche des utilisateurs StackOverflow ayant le tag COBOL ou Java dans la région Nord."""
    profiles = []
    tag = tech.lower()
    base_url = "https://api.stackexchange.com/2.3"

    try:
        # Récupérer les top répondeurs/poseurs par tag (plusieurs pages)
        for page in range(1, 4):
            url = (f"{base_url}/questions?tagged={tag}&site=stackoverflow"
                   f"&pagesize=100&page={page}&order=desc&sort=activity")
            r = requests.get(url, timeout=10)
            if r.status_code != 200:
                break
            questions = r.json().get("items", [])
            if not questions:
                break

            user_ids = list({q["owner"].get("user_id") for q in questions
                             if q.get("owner", {}).get("user_id")})
            if not user_ids:
                continue

            # Traiter par lots de 20 (limite API)
            for i in range(0, len(user_ids), 20):
                batch = user_ids[i:i+20]
                ids_str = ";".join(map(str, batch))
                users_url = f"{base_url}/users/{ids_str}?site=stackoverflow"
                ur = requests.get(users_url, timeout=10)
                if ur.status_code != 200:
                    continue

                for user in ur.json().get("items", []):
                    user_loc = (user.get("location") or "").lower()
                    if not any(kw in user_loc for kw in LOCATIONS_SO_FILTER):
                        continue
                    profiles.append({
                        "nom":          user.get("display_name", "Inconnu"),
                        "pseudo":       user.get("display_name", ""),
                        "email":        "Non public",
                        "entreprise":   "",
                        "localisation": user.get("location", ""),
                        "bio":          f"Tag: {tag} | Réputation: {user.get('reputation', 0)} | Réponses: {user.get('answer_count', 0)}",
                        "profil_url":   user.get("link", ""),
                        "repos":        user.get("answer_count", 0),
                        "followers":    user.get("reputation", 0),
                        "disponibilite": "Indéterminée",
                        "technologie":  tech,
                        "source":       "StackOverflow",
                    })
                time.sleep(0.5)
            time.sleep(1)
    except Exception as e:
        print(f"  [StackOverflow] Erreur {tech}: {e}")

    # Dédupliquer
    seen = set()
    unique = []
    for p in profiles:
        key = p["profil_url"]
        if key and key not in seen:
            seen.add(key)
            unique.append(p)
    return unique


def search_web_linkedin(tech: str) -> list[dict]:
    """Recherche de profils LinkedIn via DuckDuckGo (données publiques)."""
    if not HAS_DDG:
        return []
    profiles = []
    queries = [
        f'site:linkedin.com/in "{tech}" "Lille" -jobs -offre',
        f'site:linkedin.com/in "{tech}" "Hauts-de-France" -jobs',
        f'site:linkedin.com/in "{tech}" "Roubaix" OR "Tourcoing" OR "Valenciennes"',
        f'site:linkedin.com/in "{tech}" "Douai" OR "Dunkerque" OR "Arras" OR "Lens"',
        f'site:linkedin.com/in "{tech}" "Béthune" OR "Calais" OR "Maubeuge" OR "Cambrai"',
        f'site:linkedin.com/in "{tech}" développeur "Nord" France freelance',
    ]
    seen_urls = set()
    try:
        with DDGS() as ddgs:
            for q in queries:
                results = list(ddgs.text(q, max_results=15))
                for r in results:
                    url = r.get("href", "")
                    if "linkedin.com/in/" not in url or url in seen_urls:
                        continue
                    seen_urls.add(url)
                    title = r.get("title", "")
                    snippet = r.get("body", "")
                    # Extraire le nom depuis le titre LinkedIn
                    nom = title.split(" - ")[0].strip() if " - " in title else title
                    profiles.append({
                        "nom":          nom,
                        "pseudo":       "",
                        "email":        "Non public (LinkedIn)",
                        "entreprise":   "",
                        "localisation": "Région Lilloise",
                        "bio":          snippet[:200],
                        "profil_url":   url,
                        "repos":        0,
                        "followers":    0,
                        "disponibilite": "Open to work" if "open to work" in snippet.lower() else "Indéterminée",
                        "technologie":  tech,
                        "source":       "LinkedIn (public)",
                    })
                time.sleep(2)
    except Exception as e:
        print(f"  [LinkedIn/DDG] Erreur {tech}: {e}")
    return profiles


def search_web_forums(tech: str) -> list[dict]:
    """Recherche de profils sur dev.to, GitHub discussions, forums tech."""
    if not HAS_DDG:
        return []
    profiles = []
    queries = [
        f'site:dev.to "{tech}" "Lille" OR "Hauts-de-France" OR "Nord"',
        f'"développeur {tech}" ("Lille" OR "Valenciennes" OR "Douai" OR "Dunkerque") disponible',
        f'"développeur {tech}" ("Arras" OR "Lens" OR "Béthune" OR "Calais") freelance',
        f'"{tech} developer" "Nord-Pas-de-Calais" OR "Hauts-de-France" contact',
        f'CV "{tech}" "Nord" OR "Hauts-de-France" filetype:pdf',
        f'site:malt.fr "{tech}" "Lille" OR "Nord" freelance',
    ]
    seen_urls = set()
    try:
        with DDGS() as ddgs:
            for q in queries:
                results = list(ddgs.text(q, max_results=10))
                for r in results:
                    url = r.get("href", "")
                    if url in seen_urls or not url:
                        continue
                    seen_urls.add(url)
                    profiles.append({
                        "nom":          r.get("title", "")[:80],
                        "pseudo":       "",
                        "email":        "À vérifier",
                        "entreprise":   "",
                        "localisation": "Région Lilloise",
                        "bio":          r.get("body", "")[:200],
                        "profil_url":   url,
                        "repos":        0,
                        "followers":    0,
                        "disponibilite": "Indéterminée",
                        "technologie":  tech,
                        "source":       "Web/Forums",
                    })
                time.sleep(2)
    except Exception as e:
        print(f"  [Forums/DDG] Erreur {tech}: {e}")
    return profiles


# ══════════════════════════════════════════════════════════════════════════════
#  SCORE DE PERTINENCE
# ══════════════════════════════════════════════════════════════════════════════

def compute_score(p: dict) -> int:
    """
    Score de pertinence 0-15 (affiché /15).
    Critères : contact, disponibilité, séniorité, spécialisation, notoriété, source.
    """
    score = 0
    bio     = (p.get("bio") or "").lower()
    email   = (p.get("email") or "").lower()
    tech    = (p.get("technologie") or "").lower()
    nom     = (p.get("nom") or "").lower()
    dispo   = (p.get("disponibilite") or "").lower()
    source  = p.get("source") or ""

    # ── Contact (3 pts) ─────────────────────────────────────────────────────
    if email and "non public" not in email and "vérifier" not in email and "@" in email:
        score += 3

    # ── Disponibilité (2 pts) ────────────────────────────────────────────────
    if "open to work" in dispo or "disponible" in dispo:
        score += 2

    # ── Séniorité détectée (2 pts) ───────────────────────────────────────────
    SENIOR_KW = ["senior", "lead", "architect", "architecte", "principal",
                 "expert", "confirmé", "référent", "tech lead", "15 ans",
                 "20 ans", "10 ans", "12 ans", "expérimenté"]
    if any(k in bio or k in nom for k in SENIOR_KW):
        score += 2

    # ── Spécialisation COBOL mainframe (2 pts) ───────────────────────────────
    if tech == "cobol":
        MAINFRAME_KW = ["jcl", "cics", "vsam", "z/os", "mvs", "db2", "ibm",
                        "mainframe", "cobol", "natural", "pacbase", "ims"]
        if any(k in bio for k in MAINFRAME_KW):
            score += 2
    # ── Spécialisation Java avancée (2 pts) ──────────────────────────────────
    elif tech == "java":
        JAVA_KW = ["spring", "quarkus", "jakarta", "microservice", "kubernetes",
                   "docker", "angular", "react", "rest api", "hibernate",
                   "maven", "gradle", "jee", "j2ee", "kafka"]
        if sum(1 for k in JAVA_KW if k in bio) >= 2:
            score += 2
        elif any(k in bio for k in JAVA_KW):
            score += 1

    # ── Profil complet (1 pt) ────────────────────────────────────────────────
    if len(bio) > 50:
        score += 1

    # ── Entreprise connue (1 pt) ─────────────────────────────────────────────
    if p.get("entreprise") and len(p["entreprise"]) > 1:
        score += 1

    # ── Notoriété GitHub/SO (2 pts max) ─────────────────────────────────────
    followers = p.get("followers") or 0
    if followers >= 200:
        score += 2
    elif followers >= 30:
        score += 1

    # ── Source de qualité (1 pt) ─────────────────────────────────────────────
    if source in ("LinkedIn (public)", "Malt", "Talent.io", "APEC"):
        score += 1

    return min(score, 15)


# ══════════════════════════════════════════════════════════════════════════════
#  DÉDUPLICATION INTER-SOURCES
# ══════════════════════════════════════════════════════════════════════════════

import unicodedata
from difflib import SequenceMatcher

def _normalize_name(name: str) -> str:
    """Normalise un nom pour comparaison : minuscules, sans accents, sans ponctuation."""
    name = name.lower().strip()
    name = "".join(c for c in unicodedata.normalize("NFD", name)
                   if unicodedata.category(c) != "Mn")
    name = "".join(c if c.isalnum() or c == " " else " " for c in name)
    return " ".join(name.split())

def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio()

def deduplicate(profiles: list[dict]) -> list[dict]:
    """
    Fusionne les doublons détectés entre sources différentes.
    Critères : même nom normalisé (>88% de similarité) OU même email.
    Conserve le profil avec le score le plus élevé + complète les champs manquants.
    """
    scored = [(compute_score(p), p) for p in profiles]
    scored.sort(key=lambda x: -x[0])  # Meilleurs en premier

    kept   = []
    merged = set()

    for i, (score_i, p_i) in enumerate(scored):
        if i in merged:
            continue
        name_i  = _normalize_name(p_i.get("nom", ""))
        email_i = (p_i.get("email") or "").lower()

        for j, (score_j, p_j) in enumerate(scored):
            if j <= i or j in merged:
                continue
            name_j  = _normalize_name(p_j.get("nom", ""))
            email_j = (p_j.get("email") or "").lower()

            is_same_email = (email_i and "@" in email_i and email_i == email_j)
            is_same_name  = (name_i and len(name_i) > 4 and
                             _similarity(name_i, name_j) >= 0.88)

            if is_same_email or is_same_name:
                merged.add(j)
                # Compléter les champs vides du profil principal
                for field in ("email", "entreprise", "bio", "localisation"):
                    if not p_i.get(field) or "non public" in str(p_i.get(field,"")).lower():
                        if p_j.get(field) and "non public" not in str(p_j.get(field,"")).lower():
                            p_i[field] = p_j[field]
                # Enrichir la source
                if p_j["source"] not in p_i["source"]:
                    p_i["source"] = p_i["source"] + " + " + p_j["source"]

        kept.append(p_i)

    return kept


# ══════════════════════════════════════════════════════════════════════════════
#  ENRICHISSEMENT EMAIL (Hunter.io)
# ══════════════════════════════════════════════════════════════════════════════

def enrich_emails(profiles: list[dict], api_key: str) -> list[dict]:
    """
    Tente de trouver l'email professionnel via Hunter.io pour les profils
    qui ont une entreprise renseignée mais pas d'email public.
    Plan gratuit Hunter.io : 25 recherches/mois.
    """
    if not api_key:
        return profiles

    enriched = 0
    for p in profiles:
        email = p.get("email", "")
        if email and "@" in email and "non public" not in email.lower():
            continue  # Déjà un email
        company = p.get("entreprise", "").strip()
        nom     = p.get("nom", "").strip()
        if not company or not nom or len(nom.split()) < 2:
            continue

        parts      = nom.split()
        first_name = parts[0]
        last_name  = " ".join(parts[1:])

        try:
            url = "https://api.hunter.io/v2/email-finder"
            params = {
                "company":    company,
                "first_name": first_name,
                "last_name":  last_name,
                "api_key":    api_key,
            }
            r = requests.get(url, params=params, timeout=8)
            if r.status_code == 200:
                data = r.json().get("data", {})
                found_email = data.get("email", "")
                confidence  = data.get("score", 0)
                if found_email and confidence >= 50:
                    p["email"]        = found_email
                    p["disponibilite"] = p.get("disponibilite", "") + f" [Hunter {confidence}%]"
                    enriched += 1
            time.sleep(0.5)
        except Exception:
            pass

    print(f"  [Hunter.io] {enriched} emails enrichis")
    return profiles


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL
# ══════════════════════════════════════════════════════════════════════════════

COLUMNS = [
    ("Score ★",           9),
    ("Nom",               24),
    ("Technologie",       12),
    ("Source",            20),
    ("Email / Contact",   32),
    ("Entreprise",        22),
    ("Localisation",      22),
    ("Disponible ?",      14),
    ("Date dispo",        16),
    ("Bio / Titre",       45),
    ("URL Profil",        38),
    ("Statut recruteur",  18),
    ("✉️ Objet email",    38),
    ("✉️ Corps email",    70),
    ("Date détection",    14),
]

SOURCE_COLORS = {
    "GitHub":           COLOR_SOURCE_GH,
    "LinkedIn (public)":COLOR_SOURCE_LI,
    "StackOverflow":    COLOR_SOURCE_SO,
    "Web/Forums":       COLOR_SOURCE_WEB,
}

TECH_COLORS = {
    "COBOL": COLOR_COBOL,
    "Java":  COLOR_JAVA,
}


def hex_to_rgb(hex_color: str) -> str:
    return "FF" + hex_color.upper()


def build_excel(all_profiles: list[dict], filepath: str):
    wb = openpyxl.Workbook()

    # ── Onglet principal ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Tous les profils"
    _write_sheet(ws, all_profiles)

    # ── Onglets par technologie ───────────────────────────────────────────────
    for tech in TECHNOLOGIES:
        sub = [p for p in all_profiles if p["technologie"].upper() == tech.upper()]
        if sub:
            ws2 = wb.create_sheet(title=f"Profils {tech}")
            _write_sheet(ws2, sub)

    # ── Onglet Stats ──────────────────────────────────────────────────────────
    _write_stats(wb, all_profiles)

    wb.save(filepath)
    print(f"  ✓ Fichier sauvegardé : {filepath}")


def _write_sheet(ws, profiles: list[dict]):
    # En-tête
    header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    header_fill  = PatternFill("solid", start_color=hex_to_rgb(COLOR_HEADER))
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    wrap_align   = Alignment(wrap_text=True, vertical="top")
    thin         = Side(style="thin", color="BBBBBB")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = center_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Couleurs de score (dégradé vert→orange→rouge)
    SCORE_FILLS = {
        range(12, 16): "1E8449",  # Vert foncé (12-15)
        range(9,  12): "27AE60",  # Vert (9-11)
        range(6,   9): "F39C12",  # Orange (6-8)
        range(3,   6): "E67E22",  # Orange foncé (3-5)
        range(0,   3): "C0392B",  # Rouge (0-2)
    }
    def score_color(s):
        for r, c in SCORE_FILLS.items():
            if s in r: return c
        return "AAAAAA"

    # Données
    today = datetime.now().strftime("%Y-%m-%d")
    # Trier par score décroissant
    profiles_sorted = sorted(profiles, key=lambda p: compute_score(p), reverse=True)
    for row_idx, p in enumerate(profiles_sorted, start=2):
        tech   = p.get("technologie", "")
        source = p.get("source", "").split(" + ")[0]  # Source principale pour la couleur
        fill_color = SOURCE_COLORS.get(source, "FFFFFF")
        row_fill = PatternFill("solid", start_color=hex_to_rgb(fill_color))

        score = compute_score(p)

        # Disponibilité normalisée
        dispo_raw = (p.get("disponibilite") or "").lower()
        if "hireable" in dispo_raw or "open to work" in dispo_raw or dispo_raw == "disponible":
            dispo_label = "✅ Oui"
        elif "non disponible" in dispo_raw:
            dispo_label = "❌ Non"
        else:
            dispo_label = "❓ Inconnu"

        values = [
            score,
            p.get("nom", ""),
            tech,
            p.get("source", "").split(" + ")[0],
            p.get("email", ""),
            p.get("entreprise", ""),
            p.get("localisation", ""),
            dispo_label,
            p.get("date_dispo", ""),
            p.get("bio", ""),
            p.get("profil_url", ""),
            "À contacter",
            p.get("email_sujet", ""),
            p.get("email_corps", ""),
            today,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill   = row_fill
            cell.border = border
            cell.font   = Font(name="Arial", size=10)
            # Colonne Score : badge coloré
            if col_idx == 1:
                cell.fill      = PatternFill("solid", start_color=hex_to_rgb(score_color(score)))
                cell.font      = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx == 11 and val:  # URL Profil → hyperlien
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=10, color="1155CC", underline="single")
                cell.value = "→ Voir profil"
            elif col_idx in (13, 14):  # Colonnes email → wrap
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.font = Font(name="Arial", size=9, color="444444")
            elif col_idx == 8:  # Disponible → centré + couleur
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if "✅" in str(val):
                    cell.font = Font(name="Arial", size=10, bold=True, color="1E8449")
                elif "❌" in str(val):
                    cell.font = Font(name="Arial", size=10, color="C0392B")
                else:
                    cell.font = Font(name="Arial", size=10, color="888888")
            elif col_idx in (2, 3, 4, 6, 7, 9, 12, 13):
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = wrap_align

    # Filtre automatique
    if profiles_sorted:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(profiles_sorted)+1}"


def _write_stats(wb, all_profiles: list[dict]):
    ws = wb.create_sheet(title="Statistiques")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    header_fill = PatternFill("solid", start_color=hex_to_rgb(COLOR_HEADER))

    ws["A1"] = "Rapport de recherche – Profils COBOL & Java – Région Nord (Hauts-de-France)"
    ws["A1"].font = Font(bold=True, size=14, name="Arial", color=COLOR_HEADER)
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 28

    ws["A2"] = f"Date de génération : {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10, name="Arial")

    # Tableau récapitulatif
    headers = ["Technologie", "Source", "Nombre de profils", "Avec email public"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    row = 5
    totals = {}
    for tech in TECHNOLOGIES:
        for source in ["GitHub", "LinkedIn (public)", "StackOverflow", "Web/Forums"]:
            sub = [p for p in all_profiles
                   if p["technologie"].upper() == tech.upper() and p["source"] == source]
            with_email = sum(1 for p in sub
                             if p.get("email") and "Non public" not in p.get("email", "")
                                and "À vérifier" not in p.get("email", ""))
            if sub:
                ws.cell(row=row, column=1, value=tech)
                ws.cell(row=row, column=2, value=source)
                ws.cell(row=row, column=3, value=len(sub))
                ws.cell(row=row, column=4, value=with_email)
                for col in range(1, 5):
                    ws.cell(row=row, column=col).font = Font(name="Arial", size=10)
                row += 1
                totals[tech] = totals.get(tech, 0) + len(sub)

    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True, name="Arial")
    ws.cell(row=row, column=3, value=f"=SUM(C5:C{row-2})").font = Font(bold=True, name="Arial")

    for col in ["A", "B", "C", "D"]:
        ws.column_dimensions[col].width = 26


# ══════════════════════════════════════════════════════════════════════════════
#  DIFF & ALERTE NOUVEAUX PROFILS
# ══════════════════════════════════════════════════════════════════════════════

def load_snapshot() -> set:
    """Charge les URLs du run précédent."""
    if not os.path.exists(SNAPSHOT_FILE):
        return set()
    try:
        with open(SNAPSHOT_FILE, encoding="utf-8") as f:
            return set(json.load(f))
    except Exception:
        return set()


def save_snapshot(profiles: list[dict]):
    """Sauvegarde les URLs du run actuel pour comparaison future."""
    urls = [p.get("profil_url", "") for p in profiles if p.get("profil_url")]
    with open(SNAPSHOT_FILE, "w", encoding="utf-8") as f:
        json.dump(urls, f, ensure_ascii=False)


def find_new_profiles(profiles: list[dict], previous_urls: set) -> list[dict]:
    """Retourne uniquement les profils absents du run précédent."""
    return [p for p in profiles
            if p.get("profil_url", "") not in previous_urls]


def build_alert_report(new_profiles: list[dict], filepath: str):
    """
    Génère un fichier Excel léger avec uniquement les nouveaux profils
    et un résumé en haut de page.
    """
    if not new_profiles:
        print("  Aucun nouveau profil cette semaine.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nouveaux profils"

    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    header_fill = PatternFill("solid", start_color="FF2C3E50")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Bandeau titre
    ws.merge_cells("A1:I1")
    ws["A1"] = (f"🆕 Nouveaux profils détectés — semaine du "
                f"{datetime.now().strftime('%d/%m/%Y')} "
                f"({len(new_profiles)} profil{'s' if len(new_profiles) > 1 else ''})")
    ws["A1"].font = Font(bold=True, size=13, name="Arial", color="2C3E50")
    ws.row_dimensions[1].height = 26

    # Sous-titre par techno
    cobol_n = sum(1 for p in new_profiles if p.get("technologie","").upper() == "COBOL")
    java_n  = sum(1 for p in new_profiles if p.get("technologie","").upper() == "JAVA")
    dispo_n = sum(1 for p in new_profiles
                  if "hireable" in (p.get("disponibilite") or "").lower()
                  or "open to work" in (p.get("disponibilite") or "").lower())
    ws.merge_cells("A2:I2")
    ws["A2"] = f"COBOL : {cobol_n}   |   Java : {java_n}   |   Disponibles confirmés : {dispo_n}"
    ws["A2"].font = Font(italic=True, size=10, name="Arial", color="7F8C8D")
    ws.row_dimensions[2].height = 18

    # En-têtes colonnes
    cols = [("Score",14),("Nom",26),("Techno",10),("Source",18),
            ("Email",30),("Disponible ?",13),("Localisation",22),("Bio/Titre",40),("Profil",35)]
    for i, (name, width) in enumerate(cols, 1):
        c = ws.cell(row=3, column=i, value=name)
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A4"

    # Lignes de données — triées par score
    SCORE_COLORS = {range(10,16):"1E8449", range(6,10):"27AE60",
                    range(3,6):"F39C12", range(0,3):"C0392B"}
    def sc_color(s):
        for r,c in SCORE_COLORS.items():
            if s in r: return "FF"+c
        return "FFAAAAAA"

    # Alterner fond blanc / gris très léger
    fills = [PatternFill("solid", start_color="FFFFFFFF"),
             PatternFill("solid", start_color="FFF7F9FC")]

    sorted_new = sorted(new_profiles, key=compute_score, reverse=True)
    for row_i, p in enumerate(sorted_new, start=4):
        score = compute_score(p)
        dispo_raw = (p.get("disponibilite") or "").lower()
        dispo_label = ("✅ Oui" if "hireable" in dispo_raw or "open to work" in dispo_raw
                       else "❓ Inconnu")
        row_fill = fills[row_i % 2]
        vals = [
            score,
            p.get("nom",""),
            p.get("technologie",""),
            p.get("source","").split(" + ")[0],
            p.get("email",""),
            dispo_label,
            p.get("localisation",""),
            p.get("bio",""),
            p.get("profil_url",""),
        ]
        for col_i, val in enumerate(vals, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            if col_i == 1:  # Score badge
                cell.fill = PatternFill("solid", start_color=sc_color(score))
                cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_i == 9 and val:  # URL
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=10, color="1155CC", underline="single")
                cell.value = "→ Voir profil"
            elif col_i == 6:
                cell.alignment = Alignment(horizontal="center")
                if "✅" in str(val):
                    cell.font = Font(name="Arial", size=10, bold=True, color="1E8449")

    ws.auto_filter.ref = f"A3:{get_column_letter(len(cols))}{len(sorted_new)+3}"
    wb.save(filepath)
    print(f"  ✓ Rapport nouveaux profils : {filepath}")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Recherche profils COBOL/Java – Région Nord")
    parser.add_argument("--github-token", default="", help="Token GitHub (optionnel, augmente le rate-limit)")
    parser.add_argument("--hunter-key",  default="", help="Clé API Hunter.io (optionnel, enrichissement email)")
    parser.add_argument("--output", default=OUTPUT_FILE, help="Chemin du fichier Excel de sortie")
    args = parser.parse_args()
    # Utilise le token de la config si pas passé en argument
    if not args.github_token and GITHUB_TOKEN:
        args.github_token = GITHUB_TOKEN

    print("\n" + "═" * 60)
    print("  Recherche profils COBOL & Java – Région Nord (Hauts-de-France)")
    print(f"  Démarrage : {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print("═" * 60)

    all_profiles = []

    for tech in TECHNOLOGIES:
        print(f"\n{'─'*40}")
        print(f"  ▶ Technologie : {tech}")
        print(f"{'─'*40}")

        print("  [1/4] GitHub...")
        gh = search_github(tech, args.github_token)
        print(f"       → {len(gh)} profil(s) trouvé(s)")
        all_profiles.extend(gh)

        print("  [2/4] StackOverflow...")
        so = search_stackoverflow(tech)
        print(f"       → {len(so)} profil(s) trouvé(s)")
        all_profiles.extend(so)

        print("  [3/4] LinkedIn (recherche web publique)...")
        li = search_web_linkedin(tech)
        print(f"       → {len(li)} profil(s) trouvé(s)")
        all_profiles.extend(li)

        print("  [4/4] Forums & communautés...")
        fw = search_web_forums(tech)
        print(f"       → {len(fw)} profil(s) trouvé(s)")
        all_profiles.extend(fw)

        time.sleep(2)

    # ── Filtre qualité ────────────────────────────────────────────────────────
    before = len(all_profiles)
    all_profiles = [p for p in all_profiles if is_valid_profile(p)]
    removed = before - len(all_profiles)

    print(f"\n  [Déduplication] Fusion des doublons inter-sources...")
    before_dedup = len(all_profiles)
    all_profiles = deduplicate(all_profiles)
    deduped = before_dedup - len(all_profiles)

    print(f"  [Hunter.io] Enrichissement des emails manquants...")
    hunter_key = args.hunter_key or HUNTER_API_KEY
    all_profiles = enrich_emails(all_profiles, hunter_key)

    # ── Diff avec le run précédent ────────────────────────────────────────────
    print(f"  [Diff] Comparaison avec le snapshot précédent...")
    previous_urls  = load_snapshot()
    new_profiles   = find_new_profiles(all_profiles, previous_urls)
    save_snapshot(all_profiles)

    print(f"\n{'═'*60}")
    print(f"  Profils collectés     : {before}")
    print(f"  Filtrés (offres/excl) : {removed}")
    print(f"  Doublons fusionnés    : {deduped}")
    print(f"  Profils conservés     : {len(all_profiles)}")
    print(f"  🆕 Nouveaux ce run    : {len(new_profiles)}")
    print(f"  Génération des fichiers Excel...")

    build_excel(all_profiles, args.output)
    build_alert_report(new_profiles, NOUVEAUX_FILE)

    # Résumé alerte
    if new_profiles:
        dispo_new = [p for p in new_profiles
                     if "hireable" in (p.get("disponibilite") or "").lower()
                     or "open to work" in (p.get("disponibilite") or "").lower()]
        print(f"\n  ━━━ ALERTE NOUVEAUX PROFILS ━━━")
        print(f"  {len(new_profiles)} nouveau(x) profil(s) cette semaine")
        print(f"  dont {len(dispo_new)} disponible(s) confirmé(s)")
        top_new = sorted(new_profiles, key=compute_score, reverse=True)[:5]
        for p in top_new:
            dispo = "✅" if "hireable" in (p.get("disponibilite") or "").lower() else "  "
            print(f"  {dispo} [{compute_score(p):2}/15] {p.get('nom','')[:35]:35} | {p.get('email','Non public')[:30]}")

    print(f"\n  ✅ Terminé !")
    print(f"     Fichier principal : {args.output}")
    print(f"     Rapport nouveaux  : {NOUVEAUX_FILE}")
    print("═" * 60 + "\n")


if __name__ == "__main__":
    main()
